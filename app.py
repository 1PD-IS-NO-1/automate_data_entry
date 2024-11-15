from flask import Flask, render_template, request, jsonify, send_file
import os
import pandas as pd
import google.generativeai as genai
import json
import tempfile
from PIL import Image
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

class InvoiceProcessor:
    @staticmethod
    def convert_pdf_to_image(pdf_path):
        pdf_document = fitz.open(pdf_path)
        first_page = pdf_document.load_page(0)
        pix = first_page.get_pixmap()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
        temp_file.close()
        try:
            pix.save(temp_file.name)
        finally:
            pdf_document.close()
        return temp_file.name

    @staticmethod
    def process_file(file_path):
        file_extension = os.path.splitext(file_path.lower())[1]
        if file_extension == '.pdf':
            image_path = InvoiceProcessor.convert_pdf_to_image(file_path)
        elif file_extension in ['.png', '.jpg', '.jpeg']:
            image_path = file_path
        else:
            raise ValueError("Unsupported file type")

        # Configure Gemini API
        GOOGLE_GEMINI_API = "AIzaSyCF-jMEoZr2ji5kmJvYg4HQGWG--Bq8n84"
        genai.configure(api_key=GOOGLE_GEMINI_API)
        model = genai.GenerativeModel(model_name="gemini-1.5-flash")

        # Prepare prompts
        system_prompt = """
        You are a specialist in comprehending receipts.
        Input images in the form of receipts will be provided to you,
        and your task is to respond to questions based on the content of the input image.
        """
        user_prompt = """
        Convert Invoice data into json format with appropriate json tags as required for the data in image. 
        Data is only numbers and characters. In data there is no Special Character given. 
        If in table any row did not have 'Plate ID' of length 10 then no need to extract that data. 
        'Plate ID' always length 10 and initial 7 will be numeric and other 3 will be characters so recognize end 3 characters properly. 
        Don't extract anything except table data and Date, Party Name, Truck No. Party Name cannot be null. There will be a name on there.
        Extract also Date,Party Name,Truck No. and and give that insert that into every row end.
        """

        # Process image
        with open(image_path, "rb") as image_file:
            image_data = image_file.read()

        image_parts = [{"mime_type": "image/jpeg", "data": image_data}]
        response = model.generate_content([system_prompt, image_parts[0], user_prompt])
        
        # Clean up temporary files
        if file_extension == '.pdf':
            os.unlink(image_path)

        return response.text

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            json_response = InvoiceProcessor.process_file(filepath)
            # Clean up uploaded file
            os.unlink(filepath)
            
            # Extract JSON data
            start = json_response.find('[')
            end = json_response.rfind(']') + 1
            if start != -1 and end != -1:
                json_array = json_response[start:end]
                data = json.loads(json_array)
                return jsonify({'data': data})
            else:
                return jsonify({'error': 'Could not extract valid JSON data'}), 400
                
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/download', methods=['POST'])
def download_excel():
    try:
        data = request.json.get('data')
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Create temporary Excel file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()

        df = pd.DataFrame(data)
        wb = Workbook()
        ws = wb.active
        ws.title = "Plate Data"

        # Write data and style Excel file
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Style headers
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Apply borders and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths
        column_widths = {
            'A': 10, 'B': 20, 'C': 10, 'D': 10, 'E': 10,
            'F': 15, 'G': 20, 'H': 15, 'I': 15, 'J': 15
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(temp_file.name)
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name='invoice_data.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)